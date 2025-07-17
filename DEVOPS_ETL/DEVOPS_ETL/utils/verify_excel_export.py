"""
Script pour vérifier le contenu du fichier Excel exporté.
"""
import sys
from pathlib import Path
import openpyxl
import glob

# Pour résoudre les problèmes d'encodage sur Windows
if sys.platform.startswith('win'):
    sys.stdout.reconfigure(encoding='utf-8')
    sys.stderr.reconfigure(encoding='utf-8')

# Ajouter le répertoire racine au path pour permettre les imports relatifs
root_dir = Path(__file__).parent.parent
sys.path.append(str(root_dir))

def main():
    try:
        # Trouver le dernier fichier Excel généré
        output_dir = root_dir / "data" / "output" / "gitlab"
        excel_files = glob.glob(str(output_dir / "gitlab_users_*.xlsx"))
        
        if not excel_files:
            print("Aucun fichier Excel d'export trouvé.")
            return
        
        # Trier par date de modification (le plus récent en premier)
        latest_file = sorted(excel_files, key=lambda x: Path(x).stat().st_mtime, reverse=True)[0]
        
        print(f"Analyse du fichier Excel: {latest_file}")
        
        # Ouvrir le fichier Excel
        wb = openpyxl.load_workbook(latest_file)
        
        # Afficher les feuilles disponibles
        print(f"\nFeuilles disponibles: {wb.sheetnames}")
        
        # Vérifier le contenu de la feuille principale
        ws_users = wb["Utilisateurs GitLab"]
        
        # Compter le nombre de lignes et colonnes
        row_count = ws_users.max_row
        col_count = ws_users.max_column
        
        print(f"\nFeuille 'Utilisateurs GitLab':")
        print(f"- Nombre de lignes: {row_count}")
        print(f"- Nombre de colonnes: {col_count}")
        
        # Afficher les en-têtes
        headers = []
        for col in range(1, col_count + 1):
            cell_value = ws_users.cell(row=1, column=col).value
            headers.append(cell_value)
        
        print("\nEn-têtes des colonnes:")
        for idx, header in enumerate(headers, 1):
            print(f"  {idx}. {header}")
        
        # Afficher les premières lignes
        print("\nAperçu des données (3 premiers utilisateurs):")
        for row in range(2, min(5, row_count + 1)):
            user_data = []
            for col in range(1, col_count + 1):
                cell_value = ws_users.cell(row=row, column=col).value
                user_data.append(str(cell_value) if cell_value is not None else "")
            
            print(f"Utilisateur {row-1}:")
            for idx, (header, value) in enumerate(zip(headers, user_data)):
                if idx < 5:  # Limiter l'affichage aux 5 premières colonnes
                    print(f"  {header}: {value}")
            print("")
        
        # Vérifier le contenu de la feuille de métadonnées
        if "Métadonnées" in wb.sheetnames:
            ws_meta = wb["Métadonnées"]
            print("\nFeuille 'Métadonnées':")
            
            for row in range(1, ws_meta.max_row + 1):
                key = ws_meta.cell(row=row, column=1).value
                value = ws_meta.cell(row=row, column=2).value
                if key and value:
                    print(f"- {key}: {value}")
        
        print("\nL'export Excel a été vérifié avec succès.")
        
    except Exception as e:
        print(f"\nErreur lors de la vérification du fichier Excel: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    main()
